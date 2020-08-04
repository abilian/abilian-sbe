import hashlib
import logging
from collections import Counter
from datetime import datetime
from functools import wraps
from io import BytesIO
from operator import attrgetter
from pathlib import Path
from time import gmtime, strftime
from typing import Any, Callable, Dict, Optional, Tuple, Union

import openpyxl
import pytz
import sqlalchemy as sa
import sqlalchemy.sql.functions
from flask import current_app, flash, g, jsonify, redirect, render_template, \
    request, session, url_for
from flask.blueprints import BlueprintSetupState
from flask_login import current_user, login_required
from openpyxl.cell import WriteOnlyCell
from sqlalchemy import orm
from werkzeug.exceptions import BadRequest, InternalServerError, NotFound
from werkzeug.wrappers.response import Response
from whoosh.searching import Hit

from abilian.core.extensions import db
from abilian.core.models.subjects import Group, User, UserQuery
from abilian.core.signals import activity
from abilian.core.util import unwrap, utc_dt
from abilian.i18n import _, _l
from abilian.sbe.apps.communities.actions import register_actions
from abilian.sbe.apps.communities.blueprint import Blueprint
from abilian.sbe.apps.communities.forms import CommunityForm
from abilian.sbe.apps.communities.models import Community, Membership
from abilian.sbe.apps.communities.presenters import CommunityPresenter
from abilian.sbe.apps.communities.security import is_manager, require_admin, \
    require_manage
from abilian.sbe.apps.documents.models import Document
from abilian.services.activity import ActivityEntry
from abilian.services.security import Role
from abilian.web import csrf, views
from abilian.web.action import Endpoint
from abilian.web.nav import BreadcrumbItem
from abilian.web.views import images as image_views

__all__ = ["communities", "route", "tab"]

logger = logging.getLogger(__name__)

EPOCH = datetime.fromtimestamp(0.0, tz=pytz.utc)


def seconds_since_epoch(dt: Optional[datetime]) -> int:
    if not dt:
        return 0
    return int((utc_dt(dt) - EPOCH).total_seconds())


communities = Blueprint(
    "communities",
    __name__,
    set_community_id_prefix=False,
    template_folder="../templates",
)
route = communities.route
add_url = communities.add_url_rule
communities.record_once(register_actions)


@communities.record_once
def register_context_processors(state: BlueprintSetupState) -> None:
    @state.app.context_processor
    def communities_context_processor() -> Dict[str, Callable]:
        # helper to get an url for community image
        return {"community_image_url": image_url}


def tab(tab_name: str) -> Callable:
    """Decorator for view functions to set the current "section" this view
    belongs to."""

    def decorator(f):
        @wraps(f)
        def set_current_tab(*args: Any, **kwargs: Any) -> Union[str, Response]:
            g.current_tab = tab_name
            return f(*args, **kwargs)

        return set_current_tab

    return decorator


def default_view_kw(
    kw: Dict[str, int], obj: Any, obj_type: str, obj_id: int, **kwargs: Any
) -> Dict[str, Any]:
    """Helper for using :func:`abilian.web.views.default_view` on objects that
    belongs to a community. This function should be used as `kw_func`::

    @default_view(blueprint, Model, kw_func=default_view_kw)
    @blueprint.route("/<object_id>")
    def view():
        ...
    """
    is_community = obj_type == Community.entity_type
    community_id = kw.get("community_id")

    if is_community or community_id is None:
        # when it's a community, default_view sets community_id to 'id', we want to
        # override with the slug value.
        if obj:
            if isinstance(obj, (Hit, dict)):
                community_id = obj.get("slug" if is_community else "community_slug")
            elif is_community:
                community_id = obj.slug
            elif community_id is None and hasattr(obj, "community"):
                try:
                    community_id = obj.community.slug
                except AttributeError:
                    pass

    if community_id is not None:
        kw["community_id"] = community_id
    else:
        raise ValueError("Cannot find community_id value")

    return kw


#
# Routes
#
@route("/")
@login_required
def index() -> str:
    query = Community.query
    sort_order = request.args.get("sort", "").strip()
    if not sort_order:
        sort_order = session.get("sort_communities_order", "alpha")

    if sort_order == "activity":
        query = query.order_by(Community.last_active_at.desc())
    else:
        query = query.order_by(Community.name)

    session["sort_communities_order"] = sort_order

    if not current_user.has_role("admin"):
        # Filter with permissions
        query = query.join(Membership).filter(Membership.user == current_user)

    ctx = {"my_communities": query.all(), "sort_order": sort_order}
    return render_template("community/home.html", **ctx)


@route("/<string:community_id>/")
@views.default_view(communities, Community, "community_id", kw_func=default_view_kw)
def community() -> Response:
    return redirect(url_for("wall.index", community_id=g.community.slug))


@route("/json2")
def list_json2():
    """JSON endpoint, used for filling select boxes dynamically."""
    # TODO: make generic ?
    args = request.args

    q = args.get("q").replace("%", " ")
    if not q or len(q) < 2:
        raise BadRequest()

    query = (
        db.session.query(Community.id, Community.name)
        .filter(Community.name.ilike("%" + q + "%"))
        .distinct()
        .order_by(Community.name)
        .limit(50)
    )
    query_result = query.all()

    result = {"results": [{"id": r[0], "text": r[1]} for r in query_result]}
    return jsonify(result)


# edit views
class BaseCommunityView:
    Model = Community
    pk = "community_id"
    Form = CommunityForm
    base_template = "community/_base.html"
    decorators = [require_admin]
    view_endpoint = ""

    def init_object(self, args: Tuple[()], kwargs: Dict) -> Tuple[Tuple[()], Dict]:
        self.obj = g.community._model
        return args, kwargs

    def view_url(self) -> str:
        return url_for(self.view_endpoint, community_id=self.obj.slug)

    def get_form_kwargs(self) -> Dict[str, Any]:
        kwargs = super().get_form_kwargs()

        image = self.obj.image
        if image and "community" in g:
            image.url = image_url(self.obj, s=500)
            kwargs["image"] = image

        return kwargs


class CommunityEdit(BaseCommunityView, views.ObjectEdit):
    template = "community/edit.html"
    title = _l("Edit community")
    decorators = views.ObjectEdit.decorators + (require_admin, tab("settings"))

    def breadcrumb(self) -> BreadcrumbItem:
        url = Endpoint("communities.settings", community_id=g.community.slug)
        return BreadcrumbItem(label=_("Settings"), icon="cog", url=url)

    def before_populate_obj(self) -> None:
        form = self.form
        name = form.name.data
        if name != self.obj.name:
            self.obj.rename(name)

        del form.name

        type = form.type.data
        if type != self.obj.type:
            self.obj.type = type
            self.obj.update_roles_on_folder()
        del form.type

        self.linked_group = form.linked_group.data or None
        if self.linked_group:
            self.linked_group = Group.query.get(int(self.linked_group))
        del form.linked_group

    def after_populate_obj(self) -> None:
        self.obj.group = self.linked_group


add_url(
    "/<string:community_id>/settings",
    view_func=CommunityEdit.as_view(
        "settings",
        view_endpoint=".community",
        message_success=_l("Community settings saved successfully."),
    ),
)


class CommunityCreate(views.ObjectCreate, CommunityEdit):
    title = _l("Create community")
    decorators = views.ObjectCreate.decorators + (require_admin,)
    template = views.ObjectCreate.template
    base_template = views.ObjectCreate.base_template

    def breadcrumb(self) -> BreadcrumbItem:
        return BreadcrumbItem(label=_("Create new community"))

    def message_success(self):
        return _("Community %(name)s created successfully", name=self.obj.name)


add_url("/new", view_func=CommunityCreate.as_view("new", view_endpoint=".community"))


class CommunityDelete(BaseCommunityView, views.ObjectDelete):
    get_form_kwargs = views.ObjectDelete.get_form_kwargs


add_url(
    "/<string:community_id>/destroy",
    methods=["POST"],
    view_func=CommunityDelete.as_view(
        "delete", message_success=_l("Community destroyed.")
    ),
)

# Community Image
_DEFAULT_IMAGE = Path(__file__).parent / "data" / "community.png"
_DEFAULT_IMAGE_MD5 = hashlib.md5(_DEFAULT_IMAGE.open("rb").read()).hexdigest()
route("/_default_image")(
    image_views.StaticImageView.as_view(
        "community_default_image", set_expire=True, image=_DEFAULT_IMAGE
    )
)


class CommunityImageView(image_views.BlobView):
    id_arg = "blob_id"

    def prepare_args(self, args, kwargs):
        community = g.community
        if not community:
            raise NotFound()

        kwargs[self.id_arg] = community.image.id
        # image = open(join(dirname(__file__), "data", "community.png"), 'rb')
        return super().prepare_args(args, kwargs)


image = CommunityImageView.as_view("image", max_size=500, set_expire=True)
route("/<string:community_id>/image")(image)


def image_url(community: Union[Community, CommunityPresenter], **kwargs: Any) -> str:
    """Return proper URL for image url."""
    if not community or not community.image:
        kwargs["md5"] = _DEFAULT_IMAGE_MD5
        return url_for("communities.community_default_image", **kwargs)

    kwargs["community_id"] = community.slug
    kwargs["md5"] = community.image.md5
    return url_for("communities.image", **kwargs)


def _members_query() -> UserQuery:
    """Helper used in members views."""
    last_activity_date = sa.sql.functions.max(ActivityEntry.happened_at).label(
        "last_activity_date"
    )
    memberships = (
        User.query.options(sa.orm.undefer("photo"))
        .join(Membership)
        .outerjoin(
            ActivityEntry,
            sa.sql.and_(
                ActivityEntry.actor_id == User.id,
                ActivityEntry.target_id == Membership.community_id,
            ),
        )
        .filter(Membership.community == g.community, User.can_login == True)
        .add_columns(Membership.id, Membership.role, last_activity_date)
        .group_by(User, Membership.id, Membership.role)
        .order_by(User.last_name.asc(), User.first_name.asc())
    )

    return memberships


@route("/<string:community_id>/members")
@tab("members")
def members() -> str:
    g.breadcrumb.append(
        BreadcrumbItem(
            label=_("Members"),
            url=Endpoint("communities.members", community_id=g.community.slug),
        )
    )
    memberships = _members_query().all()
    community_threads_users = [thread.creator for thread in g.community.threads]
    threads_count = Counter(community_threads_users)

    ctx = {
        "seconds_since_epoch": seconds_since_epoch,
        "is_manager": is_manager(user=current_user),
        "memberships": memberships,
        "threads_count": threads_count,
    }
    return render_template("community/members.html", **ctx)


@route("/<string:community_id>/members", methods=["POST"])
@csrf.protect
@require_manage
def members_post() -> Response:
    community = g.community._model
    action = request.form.get("action")

    user_id = request.form.get("user")
    if not user_id:
        flash(_("You must provide a user."), "error")
        return redirect(url_for(".members", community_id=community.slug))
    user_id = int(user_id)
    user = User.query.get(user_id)

    if action in ("add-user-role", "set-user-role"):
        role = request.form["role"].lower()

        community.set_membership(user, role)

        if action == "add-user-role":
            app = unwrap(current_app)
            activity.send(app, actor=user, verb="join", object=community)

        db.session.commit()
        return redirect(url_for(".members", community_id=community.slug))

    elif action == "delete":
        membership_id = int(request.form["membership"])
        membership = Membership.query.get(membership_id)
        if membership.user_id != user_id:
            raise InternalServerError()

        community.remove_membership(user)

        app = unwrap(current_app)
        activity.send(app, actor=user, verb="leave", object=community)

        db.session.commit()
        return redirect(url_for(".members", community_id=community.slug))

    else:
        raise BadRequest(f"Unknown action: {repr(action)}")


MEMBERS_EXPORT_HEADERS = [
    _l("Name"),
    _l("email"),
    _l("Last activity in this community"),
    _l("Role"),
]

MEMBERS_EXPORT_ATTRS = ["User", "User.email", "last_activity_date", "role"]

HEADER_FONT = openpyxl.styles.Font(bold=True)
HEADER_ALIGN = openpyxl.styles.Alignment(
    horizontal="center", vertical="top", wrapText=True
)
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@route("/<string:community_id>/members/excel")
@tab("members")
def members_excel_export():
    community = g.community
    attributes = [attrgetter(a) for a in MEMBERS_EXPORT_ATTRS]
    BaseModel = db.Model
    wb = openpyxl.Workbook()

    if wb.worksheets:
        wb.remove_sheet(wb.active)

    ws_title = _("%(community)s members", community=community.name)
    ws_title = ws_title.strip()
    if len(ws_title) > 31:
        # sheet title cannot exceed 31 char. max length
        ws_title = ws_title[:30] + "…"
    ws = wb.create_sheet(title=ws_title)
    row = 0
    cells = []

    cols_width = []
    for _col, label in enumerate(MEMBERS_EXPORT_HEADERS, 1):
        value = str(label)
        cell = WriteOnlyCell(ws, value=value)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cells.append(cell)
        cols_width.append(len(value) + 1)

    ws.append(cells)

    for membership_info in _members_query().all():
        row += 1
        cells = []
        for col, getter in enumerate(attributes):
            value = None
            try:
                value = getter(membership_info)
            except AttributeError:
                pass

            if isinstance(value, (BaseModel, Role)):
                value = str(value)

            cell = WriteOnlyCell(ws, value=value)
            cells.append(value)

            # estimate width
            value = str(cell.value)
            width = max(len(l) for l in value.split("\n")) + 1
            cols_width[col] = max(width, cols_width[col])

        ws.append(cells)

    # adjust columns width
    MIN_WIDTH = 3
    MAX_WIDTH = openpyxl.utils.units.BASE_COL_WIDTH * 4

    for idx, width in enumerate(cols_width, 1):
        letter = openpyxl.utils.get_column_letter(idx)
        width = min(max(width, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[letter].width = width

    fd = BytesIO()
    wb.save(fd)
    fd.seek(0)

    response = current_app.response_class(fd, mimetype=XLSX_MIME)

    filename = "{}-members-{}.xlsx".format(
        community.slug, strftime("%d:%m:%Y-%H:%M:%S", gmtime())
    )
    response.headers["content-disposition"] = f'attachment;filename="{filename}"'

    return response


#
# Hack to redirect from urls used by the search engine.
#
@route("/doc/<int:doc_id>")
def doc(doc_id):
    doc = Document.query.get(doc_id)

    if doc is None:
        raise NotFound()

    folder = doc.parent
    while True:
        parent = folder.parent
        if parent.is_root_folder:
            break
        folder = parent
    target_community = Community.query.filter(Community.folder_id == folder.id).one()
    location = url_for(
        "documents.document_view", community_id=target_community.slug, doc_id=doc.id
    )
    return redirect(location)
